from __future__ import annotations

import ast
import io
import re
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd

MONTH_MAP = {
    "jan": 1, "january": 1, "1": 1, "1월": 1,
    "feb": 2, "february": 2, "2": 2, "2월": 2,
    "mar": 3, "march": 3, "3": 3, "3월": 3,
    "apr": 4, "april": 4, "4": 4, "4월": 4,
    "may": 5, "5": 5, "5월": 5,
    "jun": 6, "june": 6, "6": 6, "6월": 6,
    "jul": 7, "july": 7, "7": 7, "7월": 7,
    "aug": 8, "august": 8, "8": 8, "8월": 8,
    "sep": 9, "september": 9, "9": 9, "9월": 9,
    "oct": 10, "october": 10, "10": 10, "10월": 10,
    "nov": 11, "november": 11, "11": 11, "11월": 11,
    "dec": 12, "december": 12, "12": 12, "12월": 12,
}

# Metrics included in TPI matrix / displays
METRIC_ORDER = ["TPI", "P-Score", "T-Score", "T-Eng", "T-Eng.F", "T-S.B", "QR", "B.CV"]

# TPI formula alias → actual column name
ALIAS_TO_COLUMN = {
    "P":    "P-Score",
    "T":    "T-Score",
    "TENG": "T-Eng",
    "TENGF": "T-Eng.F",
    "TSB":  "T-S.B",
    "QR":   "QR",
    "BCV":  "B.CV",
}

# Canonical names for the 3 key subjects used in T-score and B.CV
_KEY_SUBJECTS = {
    "T-Eng":   {"include": ["english"], "exclude": ["found", "eng. f"]},
    "T-Eng.F": {"include": ["found", "eng. f", "eng.f", "foundation"], "exclude": []},
    "T-S.B":   {"include": ["speech"], "exclude": []},
}

REQUIRED_EXAM_COLUMNS = [
    "curriculum", "campus_type", "campus", "class_name", "student_code", "student_name",
    "exam_type", "year", "semester", "month", "subject", "item_no", "correct_answer", "student_answer",
]

# ── Risk grade info (for 지표 설명 tab) ─────────────────────────────────────
RISK_GRADE_DESCRIPTIONS = {
    "At-Risk":       "전체 TPI 하위 20% 이면서 B.CV 하위 20%",
    "High-Risk":     "전체 TPI 하위 20% (B.CV는 정상)",
    "Latent Risk":   "B.CV 하위 20% (TPI는 정상 /과목 편차가 커 잠재적 위험)",
    "Local Risk":    "캠퍼스 내 TPI 하위 20% (다른 위험등급 미해당)",
    "Top Risk":      "MAG 레벨 전용: 전체 TPI 상위 20% 이면서 B.CV 상위 20%",
    "Local Top Risk": "MAG 레벨 전용: 캠퍼스 내 TPI·B.CV 모두 상위 20% (Top Risk 미해당)",
}


# ── Basic helpers ────────────────────────────────────────────────────────────

def month_to_num(value) -> Optional[int]:
    if pd.isna(value):
        return None
    return MONTH_MAP.get(str(value).strip().lower())


def clip_0_100(series: pd.Series) -> pd.Series:
    return pd.to_numeric(series, errors="coerce").astype(float).clip(lower=0, upper=100)


def inverse_cv_score(values: pd.Series) -> float:
    """Return 100 − CV (higher = more consistent across subjects). Range [0, 100]."""
    vals = pd.to_numeric(values, errors="coerce").dropna().astype(float)
    if len(vals) <= 1:
        return 100.0
    mean = vals.mean()
    if mean == 0 or np.isnan(mean):
        return 0.0
    raw_cv = (vals.std(ddof=0) / mean) * 100
    return float(np.clip(100 - raw_cv, 0, 100))


def tscore_from_series(series: pd.Series) -> pd.Series:
    """T-Score within a group. Mean=50, SD=10, clipped to [0, 100]."""
    vals = pd.to_numeric(series, errors="coerce").astype(float)
    mean = vals.mean()
    std = vals.std(ddof=0)
    if std == 0 or np.isnan(std):
        return pd.Series(np.full(len(vals), 50.0), index=series.index)
    z = (vals - mean) / std
    return (50 + 10 * z).clip(lower=0, upper=100)


def percentile_rank(series: pd.Series) -> pd.Series:
    """Percentile rank: top student = 100%, bottom = near 0%. [0, 100]."""
    vals = pd.to_numeric(series, errors="coerce")
    return (vals.rank(method="average", pct=True) * 100).astype(float)


def normalize_exam_type(v) -> str:
    s = str(v).strip().upper()
    if "MT" in s:
        return "MT"
    if "LT" in s:
        return "LT"
    return s


def _find_subject_col(cols: List[str], include_kws: List[str], exclude_kws: List[str]) -> Optional[str]:
    """Find the first column whose lowercase name contains any include keyword but no exclude keyword."""
    for c in cols:
        cl = str(c).lower()
        if any(kw in cl for kw in include_kws) and not any(kw in cl for kw in exclude_kws):
            return c
    return None


# ── File detection / parsing ─────────────────────────────────────────────────

def detect_file_kind(name: str, expected_ext: str | None = None) -> str:
    s = Path(name).name.lower()
    ext = Path(s).suffix.lower()
    if expected_ext is not None and ext != expected_ext.lower():
        return "unknown"
    exam_markers = ["mt", "lt", "문항결과", "시험", "score", "exam"]
    student_markers = ["student", "학생", "master", "roster"]
    if ext == ".xlsx" and any(marker in s for marker in exam_markers):
        return "exam"
    if ext == ".csv" and any(marker in s for marker in student_markers):
        return "student"
    if ext == ".xlsx":
        return "exam"
    if ext == ".csv":
        return "student"
    return "unknown"


def parse_exam_filename(name: str) -> Tuple[Optional[int], Optional[int], Optional[str], Optional[str]]:
    """Parse year, month, exam_type, level from filename.
    Example: '2025년_MT_7월_GT2_문항결과.xlsx' → (2025, 7, 'MT', 'GT2')
    """
    year_m = re.search(r"(20\d{2})년", name) or re.search(r"(20\d{2})", name)
    month_m = re.search(r"(\d{1,2})월", name)
    if not month_m:
        for eng, num in [("jan", 1), ("feb", 2), ("mar", 3), ("apr", 4),
                         ("may", 5), ("jun", 6), ("jul", 7), ("aug", 8),
                         ("sep", 9), ("oct", 10), ("nov", 11), ("dec", 12)]:
            if eng in name.lower():
                month_m = type("M", (), {"group": lambda self, _=num: str(_)})()
                break
    exam_m = re.search(r"(?:^|[^a-zA-Z])(MT|LT)(?:$|[^a-zA-Z])", name, re.IGNORECASE)
    year = int(year_m.group(1)) if year_m else None
    month = int(month_m.group(1)) if month_m else None
    exam_type = exam_m.group(1).upper() if exam_m else None

    # Level: pick the first token that is not year/month/exam_type/marker
    level = None
    stem = Path(name).stem
    _skip = {"문항결과", "문항분석표", "시험", "score", "exam", "mt", "lt"}
    for token in re.split(r"[_\-\s]+", stem):
        t_lower = token.lower().strip()
        if not t_lower:
            continue
        if t_lower in _skip:
            continue
        if re.match(r"^20\d{2}(년)?$", t_lower):
            continue
        if re.match(r"^\d{1,2}월?$", t_lower):
            continue
        level = token.strip()
        break

    return year, month, exam_type, level


# ── Excel reading ────────────────────────────────────────────────────────────

def _read_excel_bytes(file_obj):
    if hasattr(file_obj, "getvalue"):
        return io.BytesIO(file_obj.getvalue())
    return file_obj


def read_single_exam(file_obj) -> pd.DataFrame:
    fname = getattr(file_obj, "name", None) or str(file_obj)
    if detect_file_kind(fname, expected_ext=".xlsx") != "exam":
        raise ValueError(f"시험 데이터 파일명이 아님: {Path(fname).name}")

    if hasattr(file_obj, "seek"):
        file_obj.seek(0)
    xls = pd.ExcelFile(_read_excel_bytes(file_obj))
    frames = []
    for sheet in xls.sheet_names:
        raw = xls.parse(sheet_name=sheet, header=None)
        if raw.shape[0] < 4:
            continue
        header = raw.iloc[2].tolist()
        df = raw.iloc[3:].copy()
        df.columns = header
        df.columns = [str(c).strip() if pd.notna(c) else c for c in df.columns]
        _EXAM_COL_ALIASES = {
            "curriculum": ["교육과정", "curriculum"],
            "campus_type": ["운영구분", "campus_type", "운영_구분"],
            "campus": ["캠퍼스", "campus", "센터"],
            "class_name": ["학급", "class_name", "class", "반"],
            "student_code": ["학번", "student_code", "학생코드"],
            "student_name": ["이름", "student_name", "학생명", "학생이름"],
            "exam_type": ["구분", "exam_type", "시험구분", "시험유형"],
            "year": ["Year", "year", "연도"],
            "semester": ["Semester", "semester", "학기"],
            "month": ["Month", "month", "월"],
            "subject": ["시험과목", "subject", "과목"],
            "item_no": ["문항 순번", "문항순번", "item_no", "문항번호"],
            "correct_answer": ["문항정답", "correct_answer", "정답"],
            "student_answer": ["학생선택", "student_answer", "학생답", "학생응답"],
        }
        _col_lower = {str(c).lower(): c for c in df.columns if pd.notna(c)}
        _rename = {}
        for internal, aliases in _EXAM_COL_ALIASES.items():
            for alias in aliases:
                if alias in df.columns:
                    _rename[alias] = internal
                    break
                elif alias.lower() in _col_lower:
                    _rename[_col_lower[alias.lower()]] = internal
                    break
        df = df.rename(columns=_rename)
        missing = [c for c in REQUIRED_EXAM_COLUMNS if c not in df.columns]
        if missing:
            raise ValueError(f"{fname} / {sheet} 필수 컬럼 누락: {missing}")
        df = df[REQUIRED_EXAM_COLUMNS].copy()
        df["student_code"] = pd.to_numeric(df["student_code"], errors="coerce").astype("Int64")
        df["year"] = pd.to_numeric(df["year"], errors="coerce").astype("Int64")
        df["item_no"] = pd.to_numeric(df["item_no"], errors="coerce").astype("Int64")
        df["month_num"] = df["month"].map(month_to_num)
        y0, m0, e0, l0 = parse_exam_filename(Path(fname).name)
        if df["year"].isna().all() and y0:
            df["year"] = y0
        if df["month_num"].isna().all() and m0:
            df["month_num"] = m0
        df["exam_type"] = df["exam_type"].fillna(e0).map(normalize_exam_type)
        df["is_correct"] = (
            df["student_answer"].astype(str).str.strip().str.lower()
            == df["correct_answer"].astype(str).str.strip().str.lower()
        ).astype(int)
        # Level: curriculum column, fallback to filename-parsed level
        df["level"] = df["curriculum"].astype(str).str.strip()
        if df["level"].eq("").all() or df["level"].eq("nan").all():
            if l0:
                df["level"] = l0
        frames.append(df)
    if not frames:
        raise ValueError(f"유효한 시트가 없음: {fname}")
    return pd.concat(frames, ignore_index=True)


# ── Statistics builders ──────────────────────────────────────────────────────

def build_item_stats(raw: pd.DataFrame) -> pd.DataFrame:
    """Item accuracy stats for the 3 key subjects only."""
    raw_3 = raw[raw["subject"].apply(_is_key_subject)].copy()
    if raw_3.empty:
        raw_3 = raw.copy()
    has_level = "level" in raw_3.columns and raw_3["level"].notna().any()
    grp = ["exam_type", "year", "month_num"] + (["level"] if has_level else []) + ["subject", "item_no"]
    item = raw_3.groupby(grp, dropna=False).agg(
        응시자수=("student_code", "nunique"),
        정답자수=("is_correct", "sum"),
    ).reset_index()
    item["문항 정답률"] = item["정답자수"] / item["응시자수"] * 100
    return item


def _is_key_subject(subj_name: str) -> bool:
    """Return True if subject name is one of the 3 key subjects."""
    sl = str(subj_name).lower()
    # Speech Building
    if "speech" in sl:
        return True
    # Eng. Foundations / English Foundation
    if any(kw in sl for kw in ["found", "eng. f", "eng.f", "foundation"]):
        return True
    # English (pure — not Foundation)
    if "english" in sl and not any(kw in sl for kw in ["found", "foundation"]):
        return True
    return False


def build_student_summary(raw: pd.DataFrame) -> pd.DataFrame:
    """
    Build per-student summary for each (exam_type, year, month_num, level).

    ALL metrics are computed exclusively from the 3 key subjects:
      English · Eng. Foundations · Speech Building
    Any other subjects (NF Studies, Listening, etc.) are IGNORED everywhere.

    Metrics:
      - P-Score  : correct / total items × 100  (3 subjects only)
      - T-Score  : standardised P-Score within group  [0, 100]
      - T-Eng    : T-Score of English subject raw score within group
      - T-Eng.F  : T-Score of Eng. Foundations raw score within group
      - T-S.B    : T-Score of Speech Building raw score within group
      - QR       : percentile rank of P-Score within group (top = 100 %)
      - B.CV     : inverse CV of the 3 subject raw scores  [0, 100]
    """
    has_level = "level" in raw.columns and raw["level"].notna().any()
    grp_exam    = ["exam_type", "year", "month_num"] + (["level"] if has_level else [])
    grp_student = grp_exam + ["campus", "campus_type", "student_code", "student_name"]

    # ── Filter to 3 key subjects ONLY ───────────────────────────────────────
    raw_3 = raw[raw["subject"].apply(_is_key_subject)].copy()
    if raw_3.empty:
        # Fallback: no matching subjects found — use all (will show empty metrics)
        raw_3 = raw.copy()

    # ── Subject raw scores (3 subjects) ─────────────────────────────────────
    subj = raw_3.groupby(grp_student + ["subject"], dropna=False).agg(
        correct=("is_correct", "sum"),
        total=("is_correct", "count"),
    ).reset_index()
    subj["subject_score"] = subj["correct"] / subj["total"] * 100
    subj_wide = subj.pivot_table(
        index=grp_student, columns="subject",
        values="subject_score", aggfunc="first",
    ).reset_index()
    subj_wide.columns.name = None

    # ── P-Score: 3 subjects only ─────────────────────────────────────────────
    overall = raw_3.groupby(grp_student, dropna=False).agg(
        total_correct=("is_correct", "sum"),
        total_items=("is_correct", "count"),
    ).reset_index()
    overall["P-Score"] = overall["total_correct"] / overall["total_items"] * 100

    score = overall.merge(subj_wide, on=grp_student, how="left")
    subject_cols = [
        c for c in score.columns
        if c not in grp_student + ["total_correct", "total_items", "P-Score"]
    ]

    # ── Identify each of the 3 key subject columns ──────────────────────────
    col_eng  = _find_subject_col(subject_cols, ["english"], ["found", "eng. f", "eng.f", "foundation"])
    col_engf = _find_subject_col(subject_cols, ["found", "eng. f", "eng.f", "foundation"], [])
    col_sb   = _find_subject_col(subject_cols, ["speech"], [])

    # ── B.CV: inverse CV of the 3 key subject raw scores ───────────────────
    bcv_src_cols = [c for c in [col_eng, col_engf, col_sb] if c is not None and c in score.columns]
    score["B.CV"] = score[bcv_src_cols].apply(inverse_cv_score, axis=1) if bcv_src_cols else 100.0

    # ── T-Score: standardised P-Score within group ───────────────────────────
    score["T-Score"] = score.groupby(grp_exam, dropna=False)["P-Score"].transform(tscore_from_series)

    # ── QR: percentile of P-Score within group ───────────────────────────────
    score["QR"] = score.groupby(grp_exam, dropna=False)["P-Score"].transform(percentile_rank)

    # ── T-Eng / T-Eng.F / T-S.B ─────────────────────────────────────────────
    for alias, col in [("T-Eng", col_eng), ("T-Eng.F", col_engf), ("T-S.B", col_sb)]:
        if col and col in score.columns:
            score[alias] = score.groupby(grp_exam, dropna=False)[col].transform(tscore_from_series)
        else:
            score[alias] = np.nan

    # ── Display column names ─────────────────────────────────────────────────
    score["캠퍼스"]   = score["campus"]
    score["학생명"]   = score["student_name"]
    score["학생코드"] = score["student_code"]
    score["월"]       = score["month_num"]
    score["시험유형"] = score["exam_type"]
    score["연도"]     = score["year"]
    if has_level:
        score["레벨"] = score["level"]

    # ── Clip & round ─────────────────────────────────────────────────────────
    metric_cols = ["P-Score", "T-Score", "T-Eng", "T-Eng.F", "T-S.B", "QR", "B.CV"]
    for c in metric_cols + subject_cols:
        if c in score.columns:
            score[c] = clip_0_100(score[c]).round(2)

    # ── Column order ─────────────────────────────────────────────────────────
    id_cols = ["캠퍼스", "학생명", "학생코드"]
    if has_level:
        id_cols.append("레벨")
    id_cols += ["시험유형", "연도", "월"]

    # Subject cols: sort so English → Eng.F → Speech Building
    def _subj_order(c: str) -> int:
        cl = c.lower()
        if "speech" in cl:
            return 2
        if any(k in cl for k in ["found", "eng. f"]):
            return 1
        return 0  # English

    subject_cols_sorted = sorted(subject_cols, key=_subj_order)

    ordered = id_cols + ["P-Score"] + subject_cols_sorted + [
        "T-Score", "T-Eng", "T-Eng.F", "T-S.B", "QR", "B.CV",
    ]
    ordered = [c for c in ordered if c in score.columns]

    return (
        score[ordered]
        .sort_values(["학생명", "학생코드", "시험유형", "월"])
        .reset_index(drop=True)
    )


# ── TPI formula engine ───────────────────────────────────────────────────────

class SafeFormulaEvaluator(ast.NodeVisitor):
    allowed_nodes = (
        ast.Expression, ast.BinOp, ast.UnaryOp, ast.Load,
        ast.Name, ast.Constant,
        ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow,
        ast.USub, ast.UAdd,
    )

    def __init__(self, variables: Dict[str, float]):
        self.variables = variables

    def visit(self, node):
        if not isinstance(node, self.allowed_nodes):
            raise ValueError("허용되지 않은 수식 요소가 포함되어 있습니다.")
        return super().visit(node)

    def visit_Expression(self, node):
        return self.visit(node.body)

    def visit_Name(self, node):
        if node.id not in self.variables:
            raise ValueError(f"알 수 없는 변수: {node.id}")
        return float(self.variables[node.id])

    def visit_Constant(self, node):
        if isinstance(node.value, (int, float)):
            return float(node.value)
        raise ValueError("숫자만 사용할 수 있습니다.")

    def visit_UnaryOp(self, node):
        val = self.visit(node.operand)
        if isinstance(node.op, ast.USub):
            return -val
        if isinstance(node.op, ast.UAdd):
            return val
        raise ValueError("허용되지 않은 단항 연산자")

    def visit_BinOp(self, node):
        left = self.visit(node.left)
        right = self.visit(node.right)
        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            return 0.0 if right == 0 else left / right
        if isinstance(node.op, ast.Pow):
            return left ** right
        raise ValueError("허용되지 않은 연산자")


def apply_tpi_formula(df: pd.DataFrame, formula: str) -> pd.DataFrame:
    out = df.copy()
    aliases = {}
    for alias, col in ALIAS_TO_COLUMN.items():
        aliases[alias] = (
            out[col].fillna(0).astype(float)
            if col in out.columns
            else pd.Series(np.zeros(len(out)), index=out.index)
        )
    tree = ast.parse(formula, mode="eval")
    values = []
    for idx in out.index:
        vars_i = {k: float(v.loc[idx]) for k, v in aliases.items()}
        values.append(SafeFormulaEvaluator(vars_i).visit(tree))
    out["TPI"] = pd.Series(values, index=out.index).clip(lower=0, upper=100).round(2)
    return out


def make_default_formula(enabled_weights: Dict[str, float]) -> str:
    parts = []
    total = 0.0
    for alias, w in enabled_weights.items():
        if w and w > 0:
            parts.append(f"({alias}*{w})")
            total += w
    return f"({' + '.join(parts)}) / {total}" if parts else "0"


# ── Risk grade computation ───────────────────────────────────────────────────

def compute_risk_grades(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add TPI rank, percentile, risk grade, and reason columns.

    Risk evaluation order (within same 시험유형 · 월 · 레벨 group):

    For MAG-level students (level starts with MAG):
      1. Top Risk      /TPI ≥ 80th pct  AND B.CV ≥ 80th pct  (globally)
      2. Local Top Risk— campus TPI ≥ 80th AND campus B.CV ≥ 80th (but NOT Top Risk)
      Then fall through to normal risk levels below.

    For all students:
      3. At-Risk       /TPI ≤ 20th pct  AND B.CV ≤ 20th pct
      4. High-Risk     /TPI ≤ 20th pct  (B.CV normal)
      5. Latent Risk   /B.CV ≤ 20th pct (TPI normal)
      6. Local Risk    /campus TPI ≤ 20th pct (none of the above)
    """
    out = df.copy()
    if "TPI" not in out.columns:
        return out

    has_level = "레벨" in out.columns
    grp_keys  = ["시험유형", "월"] + (["레벨"] if has_level else [])
    campus_keys = grp_keys + ["캠퍼스"]

    # ── Ranks ──────────────────────────────────────────────────────────────
    out["TPI랭크(전체)"] = (
        out.groupby(grp_keys, dropna=False)["TPI"]
        .rank(ascending=False, method="min")
        .astype("Int64")
    )
    out["TPI랭크(캠퍼스)"] = (
        out.groupby(campus_keys, dropna=False)["TPI"]
        .rank(ascending=False, method="min")
        .astype("Int64")
    )

    # ── Percentiles ─────────────────────────────────────────────────────────
    # TPI분위: 1등 = 100, 꼴찌 = near 0
    out["TPI분위"] = (
        out.groupby(grp_keys, dropna=False)["TPI"]
        .transform(percentile_rank)
        .round(1)
    )

    has_bcv = "B.CV" in out.columns
    if has_bcv:
        out["_bcv_pct"]        = out.groupby(grp_keys,    dropna=False)["B.CV"].transform(percentile_rank)
        out["_campus_bcv_pct"] = out.groupby(campus_keys, dropna=False)["B.CV"].transform(percentile_rank)
    else:
        out["_bcv_pct"]        = 50.0
        out["_campus_bcv_pct"] = 50.0

    out["_campus_tpi_pct"] = (
        out.groupby(campus_keys, dropna=False)["TPI"]
        .transform(percentile_rank)
    )

    # ── Row-wise risk assignment ─────────────────────────────────────────────
    def _assign(row):
        tpi_pct        = float(row.get("TPI분위",          50) or 50)
        bcv_pct        = float(row.get("_bcv_pct",         50) or 50)
        c_tpi_pct      = float(row.get("_campus_tpi_pct",  50) or 50)
        c_bcv_pct      = float(row.get("_campus_bcv_pct",  50) or 50)
        tpi_val        = float(row.get("TPI",   0) or 0)
        bcv_val        = float(row.get("B.CV",  0) or 0)
        level          = str(row.get("레벨", "") or "")
        is_mag         = bool(re.match(r"(?i)^MAG\d*", level.strip()))

        tpi_low  = tpi_pct   <= 20
        bcv_low  = bcv_pct   <= 20
        c_tpi_lo = c_tpi_pct <= 20
        tpi_high = tpi_pct   >= 80
        bcv_high = bcv_pct   >= 80
        c_tpi_hi = c_tpi_pct >= 80
        c_bcv_hi = c_bcv_pct >= 80

        # ── MAG top performer labels ────────────────────────────────────────
        if is_mag:
            if tpi_high and bcv_high:
                reason = (
                    f"MAG 레벨: TPI {tpi_pct:.0f}분위(상위 20%), "
                    f"B.CV {bcv_pct:.0f}분위(상위 20%) /TPI {tpi_val:.1f}, B.CV {bcv_val:.1f}"
                )
                return "Top Risk", reason
            if c_tpi_hi and c_bcv_hi:
                reason = (
                    f"MAG 레벨: 캠퍼스 내 TPI 상위 20%({c_tpi_pct:.0f}분위), "
                    f"캠퍼스 내 B.CV 상위 20%({c_bcv_pct:.0f}분위) /TPI {tpi_val:.1f}, B.CV {bcv_val:.1f}"
                )
                return "Local Top Risk", reason

        # ── Standard risk levels ────────────────────────────────────────────
        if tpi_low and bcv_low:
            reason = (
                f"TPI {tpi_pct:.0f}분위(전체 하위 20%), "
                f"B.CV {bcv_pct:.0f}분위(하위 20%) /TPI {tpi_val:.1f}, B.CV {bcv_val:.1f}"
            )
            return "At-Risk", reason

        if tpi_low:
            reason = (
                f"TPI {tpi_pct:.0f}분위(전체 하위 20%) /TPI {tpi_val:.1f} "
                f"(B.CV {bcv_val:.1f} 정상)"
            )
            return "High-Risk", reason

        if bcv_low:
            reason = (
                f"B.CV {bcv_pct:.0f}분위(하위 20%) /B.CV {bcv_val:.1f}, "
                f"과목 간 편차 과대 (TPI {tpi_val:.1f} 정상)"
            )
            return "Latent Risk", reason

        if c_tpi_lo:
            reason = (
                f"캠퍼스 내 TPI {c_tpi_pct:.0f}분위(하위 20%) /TPI {tpi_val:.1f}"
            )
            return "Local Risk", reason

        return "", ""

    result = out.apply(_assign, axis=1, result_type="expand")
    result.columns = ["위험등급", "사유"]
    out["위험등급"] = result["위험등급"]
    out["사유"]     = result["사유"]

    # Clean temp columns
    out = out.drop(columns=["_bcv_pct", "_campus_bcv_pct", "_campus_tpi_pct"], errors="ignore")
    return out


# ── TPI matrix (pivot) ────────────────────────────────────────────────────────

def period_sort_key(label: str) -> Tuple[int, int]:
    m = re.match(r"(MT|LT)-(\d+)월", label)
    if not m:
        return (99, 99)
    return (0 if m.group(1) == "MT" else 1, int(m.group(2)))


def build_tpi_matrix(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["캠퍼스", "학생명", "학생코드", "레벨", "지표"])
    work = df.copy()
    has_level = "레벨" in work.columns
    work["기간"] = work["시험유형"].astype(str) + "-" + work["월"].astype(int).astype(str) + "월"

    id_cols = ["캠퍼스", "학생명", "학생코드"]
    if has_level:
        id_cols.append("레벨")
    id_cols.append("기간")

    long_parts = []
    for metric_name in METRIC_ORDER:
        if metric_name not in work.columns:
            continue
        tmp = work[id_cols + [metric_name]].copy().rename(columns={metric_name: "값"})
        tmp["지표"] = metric_name
        long_parts.append(tmp)

    if not long_parts:
        return pd.DataFrame()
    long_df = pd.concat(long_parts, ignore_index=True)

    pivot_idx = ["캠퍼스", "학생명", "학생코드"]
    if has_level:
        pivot_idx.append("레벨")
    pivot_idx.append("지표")

    pivot = (
        long_df
        .pivot_table(index=pivot_idx, columns="기간", values="값", aggfunc="first")
        .reset_index()
    )
    period_cols = sorted(
        [c for c in pivot.columns if c not in pivot_idx],
        key=period_sort_key,
    )
    pivot = pivot[pivot_idx + period_cols].copy()
    metric_order_map = {m: i for i, m in enumerate(METRIC_ORDER)}
    pivot["_ord"] = pivot["지표"].map(metric_order_map).fillna(999)
    pivot = (
        pivot
        .sort_values(["학생명", "학생코드", "_ord"])
        .drop(columns=["_ord"])
        .reset_index(drop=True)
    )
    for c in period_cols:
        pivot[c] = pd.to_numeric(pivot[c], errors="coerce").round(2)
    return pivot


# ── Export helpers ────────────────────────────────────────────────────────────

def to_csv_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_csv(buf, index=False, encoding="utf-8-sig")
    return buf.getvalue()


def to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "결과") -> bytes:
    """Export DataFrame to xlsx bytes with header styling and auto-column widths."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]

        # Header style
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin = Side(style="thin", color="BBBBBB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = header_align
            cell.border = border

        # Auto column widths
        for col_idx, col_name in enumerate(df.columns, 1):
            values_len = (
                df[col_name].astype(str).str.len().max()
                if len(df) > 0 else 0
            )
            width = min(max(len(str(col_name)) + 2, int(values_len) + 2), 40)
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

    return buf.getvalue()
